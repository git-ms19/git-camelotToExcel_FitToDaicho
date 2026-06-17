import importlib.util
import os
import sys
import tempfile
import types
import unittest
from unittest.mock import patch

import pandas as pd
from openpyxl import load_workbook


SCRIPT_PATH = os.path.join(
    os.path.dirname(__file__), "camelotToExcel_FitToDaicho2.0.py"
)


def load_script_functions():
    with open(SCRIPT_PATH, encoding="utf-8") as source_file:
        source = source_file.read()
    source = source.split("\nroot = tk.Tk()", 1)[0]
    spec = importlib.util.spec_from_loader("camelot_app_for_test", loader=None)
    module = importlib.util.module_from_spec(spec)
    module.__file__ = SCRIPT_PATH
    sys.modules[spec.name] = module
    sys.modules.setdefault("camelot", types.SimpleNamespace())
    exec(compile(source, SCRIPT_PATH, "exec"), module.__dict__)
    return module


app = load_script_functions()


class FakeTextContainer(app.LTTextContainer):
    def __init__(self, text):
        self.text = text

    def get_text(self):
        return self.text


class PageOutputNameTests(unittest.TestCase):
    def test_extract_page_names_normalizes_full_width_digits(self):
        pages = [
            [FakeTextContainer("\u5317\u5206\u4f1a\n\uff10\uff11\u73ed")],
            [FakeTextContainer("\u97f3\u7fbd\u5c71\u5206\u4f1a 05\u73ed")],
        ]
        with patch.object(app, "extract_pages", return_value=pages):
            names = app.extract_page_names("input.pdf")

        self.assertEqual({1: "\u531701", 2: "\u97f3\u7fbd\u5c7105"}, names)

    def test_duplicate_names_receive_sequence_on_every_page(self):
        tables = [
            app.ExtractedTable(pd.DataFrame([[1]]), 1, 99.0, "\u531701"),
            app.ExtractedTable(pd.DataFrame([[2]]), 2, 99.0, "\u531701"),
            app.ExtractedTable(
                pd.DataFrame([[3]]), 3, 99.0, "\u97f3\u7fbd\u5c7105"
            ),
        ]

        self.assertEqual(
            ["\u531701p1", "\u531701p2", "\u97f3\u7fbd\u5c7105"],
            app.make_output_names(tables),
        )

    def test_excel_names_keep_suffix_within_31_characters(self):
        long_name = "\u5317" * 31
        tables = [
            app.ExtractedTable(pd.DataFrame([[1]]), 1, 99.0, long_name),
            app.ExtractedTable(pd.DataFrame([[2]]), 2, 99.0, long_name),
        ]

        names = app.make_output_names(tables, max_length=31)

        self.assertEqual(["p1", "p2"], [name[-2:] for name in names])
        self.assertTrue(all(len(name) == 31 for name in names))

    def test_csv_json_and_excel_use_page_names(self):
        tables = [
            app.ExtractedTable(pd.DataFrame([["a"]]), 1, 99.0, "\u531701"),
            app.ExtractedTable(pd.DataFrame([["b"]]), 2, 99.0, "\u531701"),
        ]
        with tempfile.TemporaryDirectory() as output_folder:
            csv_paths = app.export_tables(
                tables, "input.pdf", output_folder, "csv"
            )
            json_paths = app.export_tables(
                tables, "input.pdf", output_folder, "json"
            )
            excel_path = app.export_tables(
                tables, "input.pdf", output_folder, "excel"
            )[0]

            self.assertEqual(
                ["\u531701p1.csv", "\u531701p2.csv"],
                [os.path.basename(path) for path in csv_paths],
            )
            self.assertEqual(
                ["\u531701p1.json", "\u531701p2.json"],
                [os.path.basename(path) for path in json_paths],
            )
            workbook = load_workbook(excel_path, read_only=True)
            self.assertEqual(["\u531701p1", "\u531701p2"], workbook.sheetnames)
            workbook.close()


if __name__ == "__main__":
    unittest.main()
